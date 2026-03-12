import os
import time
from datetime import datetime

def generate_financial_report(data: dict, output_filename: str) -> str:
    """
    Transforms the validated CFO JSON into a 3-sheet Enterprise Excel Dashboard.
    Features: KPIs, Programmatic Charts, Conditional Matrix, Audit Traceability.
    """
    base_dir = os.path.join(os.getcwd(), "data", "reports")
    os.makedirs(base_dir, exist_ok=True)
    full_path = os.path.join(base_dir, output_filename)
    
    print(f"[EXCEL SKILL] Starting Enterprise Generation via OpenPyXL: {full_path}")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference, Series
    except ImportError:
        print("[EXCEL SKILL] Error: openpyxl module not found.")
        return None

    try:
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        
        cfo = data.get("Financial_Report", {})
        if not cfo:
            print("[EXCEL SKILL] No 'Financial_Report' key found in data.")
            return None
            
        # Common Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        center_align = Alignment(horizontal="center", vertical="center")
        wrap_align = Alignment(wrap_text=True, vertical="top")
        bold_font = Font(bold=True)
        
        # -------------------------------------------------------------------
        # SHEET 1: EXECUTIVE DASHBOARD
        # -------------------------------------------------------------------
        ws1 = wb.create_sheet(title="Executive Dashboard")
        narrative = cfo.get("narrative", {})
        conf_score = cfo.get("confidence_score", 0)
        
        ws1.column_dimensions['A'].width = 3
        ws1.column_dimensions['B'].width = 30
        ws1.column_dimensions['C'].width = 80
        
        # Title
        ws1.merge_cells('B2:C2')
        ws1['B2'] = "NVIDIA: ENTERPRISE CFO FINANCIAL COHERENCE AUDIT"
        ws1['B2'].font = Font(bold=True, size=16, color="1F4E78")
        
        # Confidence Score
        ws1['B4'] = "Audit Confidence Score:"
        ws1['B4'].font = bold_font
        ws1['C4'] = f"{conf_score} / 100"
        
        score_color = "00B050" if conf_score >= 80 else ("FFC000" if conf_score >= 50 else "FF0000")
        ws1['C4'].font = Font(bold=True, color=score_color, size=14)
        
        if conf_score < 80:
            ws1.merge_cells('B5:C5')
            ws1['B5'] = "⚠️ WARNING: Human Audit Recommended due to Adversarial Rejections or Missing Data."
            ws1['B5'].font = Font(bold=True, color="FF0000")
            
        # Narrative Sections
        row_cursor = 7
        sections = [
            ("Financial Performance Summary", narrative.get("performance_summary", "")),
            ("Efficiency & Cost Structure", narrative.get("efficiency_assessment", "")),
            ("Risk Signals", narrative.get("risk_signals", "")),
            ("Strategic Outlook", narrative.get("strategic_outlook", ""))
        ]
        
        for title, content in sections:
            ws1.cell(row=row_cursor, column=2, value=title).font = bold_font
            ws1.cell(row=row_cursor, column=2).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            ws1.cell(row=row_cursor+1, column=2).value = "Details:"
            
            # Format bullet points if they exist, or just wrap
            cell_content = ws1.cell(row=row_cursor+1, column=3, value=content)
            cell_content.alignment = wrap_align
            ws1.row_dimensions[row_cursor+1].height = 60 # Taller for bullets
            
            row_cursor += 3

        # -------------------------------------------------------------------
        # NEW: PERFORMANCE AUDIT (Framework Demo)
        # -------------------------------------------------------------------
        metadata = data.get("Metadata", {})
        if metadata:
            ws1.cell(row=row_cursor, column=2, value="Psiquis-X: Performance & Framework Audit").font = bold_font
            ws1.cell(row=row_cursor, column=2).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            total_lat = metadata.get("total_latency", 0)
            ws1.cell(row=row_cursor+1, column=2).value = "Total Pipeline Latency:"
            ws1.cell(row=row_cursor+1, column=3).value = f"{total_lat:.2f} seconds"
            
            # Simplified node summary
            node_stats = metadata.get("node_stats", [])
            summary_parts = []
            for ns in node_stats:
                summary_parts.append(f"{ns['node'].upper()} ({ns.get('latency', 0):.1f}s)")
            
            ws1.cell(row=row_cursor+2, column=2).value = "Engine Lifecycle (Nodes):"
            ws1.cell(row=row_cursor+2, column=3).value = " -> ".join(summary_parts)
            ws1.cell(row=row_cursor+2, column=3).font = Font(size=9, italic=True)
            
            row_cursor += 4

        # Prepare Chart Data (Hidden area in Dashboard)
        # We need to extract Revenue vs Opex per Period
        raw_metrics_for_chart = cfo.get("metrics", [])
        periods = []
        rev_map = {}
        opex_map = {}
        
        for m in raw_metrics_for_chart:
            fy = m.get("fiscal_year", "Unknown")
            name = m.get("metric_name", "").lower()
            val = m.get("value_reported", 0)
            
            if fy not in periods:
                periods.append(fy)
                rev_map[fy] = 0
                opex_map[fy] = 0
                
            if "revenue" in name:
                rev_map[fy] = val
            elif "operating exp" in name:
                opex_map[fy] = val
                
        periods = sorted(periods) # E.g. FY24, FY25, FY26
        
        chart_start_col = 10 # Column J
        ws1.cell(row=2, column=chart_start_col, value="Period")
        ws1.cell(row=2, column=chart_start_col+1, value="Total Revenue")
        ws1.cell(row=2, column=chart_start_col+2, value="Operating Expenses")
        
        chart_end_row = 2
        for p in periods:
            chart_end_row += 1
            ws1.cell(row=chart_end_row, column=chart_start_col, value=p)
            ws1.cell(row=chart_end_row, column=chart_start_col+1, value=rev_map[p])
            ws1.cell(row=chart_end_row, column=chart_start_col+2, value=opex_map[p])
            
        # Draw Bar Chart
        if len(periods) > 0:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Revenue vs Operating Expenses (Billions)"
            chart.y_axis.title = 'Billions USD'
            chart.x_axis.title = 'Fiscal Period'
            
            data_ref = Reference(ws1, min_col=chart_start_col+1, min_row=2, max_col=chart_start_col+2, max_row=chart_end_row)
            cats_ref = Reference(ws1, min_col=chart_start_col, min_row=3, max_row=chart_end_row)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.shape = 4
            
            ws1.add_chart(chart, "E7")

        # -------------------------------------------------------------------
        # SHEET 2: COHORT ANALYSIS (Variance Matrix)
        # -------------------------------------------------------------------
        ws2 = wb.create_sheet(title="Cohort Analysis (Variance)")
        variance_data = cfo.get("variance_analysis", [])
        
        headers2 = ["Metric Name", "Comparison Period", "Variance Growth %", "Coherence Flag"]
        for col_idx, hdr in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_idx, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        red_fill = PatternFill(start_color="FFC7CE", end_color="9C0006", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="006100", fill_type="solid")
        amber_fill = PatternFill(start_color="FFEB9C", end_color="9C5700", fill_type="solid")
            
        for r_idx, row_dict in enumerate(variance_data, 2):
            ws2.cell(row=r_idx, column=1, value=row_dict.get("metric_name"))
            ws2.cell(row=r_idx, column=2, value=row_dict.get("comparison_period")).alignment = center_align
            
            growth_cell = ws2.cell(row=r_idx, column=3, value=row_dict.get("growth_percentage", 0) / 100.0)
            growth_cell.number_format = "0.0%"
            
            flag = row_dict.get("coherence_flag", "")
            flag_cell = ws2.cell(row=r_idx, column=4, value=flag)
            
            # Conditional Application
            if "Risk" in flag or "Compression" in flag or "Deterioration" in flag:
                flag_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                flag_cell.font = Font(color="9C0006", bold=True)
            elif "Positive" in flag or "Leverage" in flag or "OK" in flag:
                flag_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                flag_cell.font = Font(color="006100", bold=True)
            else:
                flag_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                flag_cell.font = Font(color="9C5700", bold=True)

        for col_idx in range(1, 5):
            ws2.column_dimensions[get_column_letter(col_idx)].width = 30

        # -------------------------------------------------------------------
        # SHEET 3: RAW DATA TRACEABILITY
        # -------------------------------------------------------------------
        ws3 = wb.create_sheet(title="Raw Data Traceability")
        raw_metrics = cfo.get("metrics", [])
        
        headers3 = ["Fiscal Year", "Metric Name", "Value (Billions)", "GAAP Standard", "Source File", "Page Reference", "Literal Snippet"]
        for col_idx, hdr in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col_idx, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        for r_idx, m in enumerate(raw_metrics, 2):
            ws3.cell(row=r_idx, column=1, value=m.get("fiscal_year")).alignment = center_align
            ws3.cell(row=r_idx, column=2, value=m.get("metric_name"))
            
            val_cell = ws3.cell(row=r_idx, column=3, value=m.get("value_reported"))
            val_cell.number_format = "_($* #,##0.0_);_($* (#,##0.0);_($* \"-\"??_);_(@_)"
            
            ws3.cell(row=r_idx, column=4, value="GAAP" if m.get("is_gaap") else "Non-GAAP").alignment = center_align
            
            source_file = m.get("source_file", "Unknown")
            ws3.cell(row=r_idx, column=5, value=source_file).alignment = center_align
            
            page_ref = m.get("page_reference", "Unknown")
            # Build absolute URI for Hyperlink if environment allows
            # Assuming files are in current working directory
            cwd = os.getcwd()
            file_uri = f"file:///{cwd}/{source_file}".replace("\\", "/")
            
            page_cell = ws3.cell(row=r_idx, column=6, value=page_ref)
            page_cell.hyperlink = file_uri
            page_cell.font = Font(color="0000FF", underline="single")
            page_cell.alignment = center_align
            
            snippet_cell = ws3.cell(row=r_idx, column=7, value=m.get("literal_snippet", ""))
            snippet_cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws3.row_dimensions[r_idx].height = 45 # Make room for snippet
            
        ws3.column_dimensions['A'].width = 15
        ws3.column_dimensions['B'].width = 25
        ws3.column_dimensions['C'].width = 20
        ws3.column_dimensions['D'].width = 20
        ws3.column_dimensions['E'].width = 40
        ws3.column_dimensions['F'].width = 25
        ws3.column_dimensions['G'].width = 80

        # Clean up
        if default_sheet in wb:
            wb.remove(default_sheet)
        wb.active = 0 # Focus on Executive Dashboard
        
        wb.save(full_path)
        print(f"[EXCEL SKILL] ✅ Success! Enterprise CFO Dashboard saved natively to {full_path}")
        
        try: os.startfile(full_path)
        except: pass
            
        return full_path

    except Exception as e:
        print(f"[EXCEL SKILL] Error generating CFO Excel: {e}")
        import traceback
        traceback.print_exc()
        return None
